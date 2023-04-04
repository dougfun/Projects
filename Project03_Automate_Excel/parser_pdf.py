import os
import openpyxl as xl
import pdfplumber

# Searching for pdf files:

for pdf_file in os.listdir('pdfs'):

    if pdf_file.lower().endswith('.pdf'):
        
        try:
            # Openning the pdf file
            excel = xl.load_workbook(r'C:\Users\dougi\Desktop\Project03\database_inspections.xlsx')
            tab = excel.active
            initial_line = len(tab['A']) + 1

            # Reading pdf and extract the data
            pdf = pdfplumber.open(f'pdfs\\{pdf_file}')
            page = pdf.pages[0] # Our pdf onyl has one page
            pdf_data = page.extract_table()

            for index, data in enumerate(pdf_data[1:], start = initial_line):
                
                if data[0] == '':
                    pass
                else:
                    tab.cell(row = index, column = 1).value = data[0]
                    tab.cell(row = index, column = 2).value = data[1]
                    tab.cell(row = index, column = 3).value = data[2]
                    tab.cell(row = index, column = 4).value = data[3]
                    tab.cell(row = index, column = 5).value = data[4]
            
            pdf.close()
            excel.save(r'C:\Users\dougi\Desktop\Project03\database_inspections.xlsx')
            excel.close()

        except Exception as e:
            with open('log_errors.txt', 'a') as log:
                log.write(f'Something happend while trying to extract information from file {pdf_file}')
                log.write(f'Error: {e}')
    else:
        with open('log_errors.txt', 'a') as log:
            log.write(f'The file {pdf_file} is not valid!\n')